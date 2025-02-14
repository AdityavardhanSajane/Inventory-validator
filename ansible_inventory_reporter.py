from typing import Optional, List, Dict, Any, cast, Literal, Union
import sys
import os
import re
from datetime import datetime
import requests
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, TaskID
from rich.prompt import Prompt, Confirm
import logging
from config import TOWER_BASE_URLS, XLR_CONFIG, AUTH_CONFIG
from logger import setup_logger
from utils import sanitize_filename, validate_spk_name, ensure_output_directory
from urllib.parse import urlparse
from xlr_client import XLRClient
from credential_manager import CredentialManager

console = Console()
logger = setup_logger()

class AnsibleTowerInventoryReporter:
    def __init__(self):
        """Initialize the reporter with required attributes"""
        self.session = requests.Session()
        self.console = Console()
        self.environments: Dict[str, List[str]] = {
            'NON_PROD': ['DEV', 'LLE'],
            'PROD': ['PROD']
        }
        self.output_dir = ensure_output_directory()
        self.xlr_client = None
        self.credential_manager = CredentialManager()
        self.ansible_instance: Optional[Literal['NON_PROD', 'PROD']] = None
        self.username: Optional[str] = None
        self.password: Optional[str] = None
        self._progress: Optional[Progress] = None
        self._current_task: Optional[TaskID] = None

    def _start_progress(self, description: str) -> None:
        """Start a new progress indicator with enhanced visuals"""
        if self._progress is None:
            self._progress = Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
                console=self.console
            )
            self._progress.start()
        self._current_task = self._progress.add_task(
            description,
            total=100  # Now using percentage-based progress
        )

    def _update_progress(self, description: str, advance: int = 10) -> None:
        """Update the current progress indicator with percentage advancement"""
        if self._progress is not None and self._current_task is not None:
            self._progress.update(
                self._current_task,
                description=description,
                advance=advance,
                refresh=True
            )

    def _stop_progress(self) -> None:
        """Stop and cleanup progress indicator"""
        if self._progress is not None:
            if self._current_task is not None:
                self._progress.remove_task(self._current_task)
            self._progress.stop()
            self._progress = None
            self._current_task = None

    def _validate_nbk_id(self, nbk_id: str) -> bool:
        """Validate NBK ID format"""
        if not nbk_id or len(nbk_id) < 3:
            return False
        # NBK ID should be alphanumeric and at least 3 characters
        return bool(re.match(r'^[A-Za-z0-9]{3,}$', nbk_id.strip()))

    def _get_new_credentials(self) -> bool:
        """Get new credentials from user and save them"""
        try:
            # Stop any existing progress indicators before user input
            self._stop_progress()

            while True:
                try:
                    nbk_id = Prompt.ask("[cyan]Enter your NBK ID[/cyan]")
                    if self._validate_nbk_id(nbk_id):
                        break
                    self.console.print(
                        "[red]Invalid NBK ID format.[/red]\n"
                        "[yellow]NBK ID should:\n"
                        "1. Be at least 3 characters long\n"
                        "2. Contain only letters and numbers[/yellow]"
                    )
                except (EOFError, KeyboardInterrupt):
                    logger.error("Authentication cancelled by user")
                    self.console.print("\n[yellow]Authentication cancelled by user.[/yellow]")
                    return False

            self.username = nbk_id

            max_retries = 3
            for attempt in range(max_retries):
                try:
                    self.password = Prompt.ask(
                        "[cyan]Enter your password[/cyan] (password will be hidden)",
                        password=True
                    )
                    if self.password:  # Ensure we got a non-empty password
                        break
                    self.console.print("[red]Password cannot be empty.[/red]")
                except (EOFError, KeyboardInterrupt):
                    if attempt < max_retries - 1:
                        self.console.print("[yellow]Password entry cancelled. Please try again.[/yellow]")
                        continue
                    logger.error("Password entry cancelled after maximum attempts")
                    self.console.print("[red]Authentication cancelled after maximum attempts.[/red]")
                    return False

            if not self.username or not self.password:
                logger.error("Missing credentials after input")
                return False

            # Start progress for credential saving
            self._start_progress("[cyan]Saving credentials securely...[/cyan]")

            # Save credentials
            if self.credential_manager.save_credentials(self.username, self.password):
                self._update_progress("[green]Credentials saved successfully![/green]")
                logger.info("New credentials saved successfully")
            else:
                self._update_progress("[yellow]Warning: Unable to save credentials securely.[/yellow]")
                logger.warning("Failed to save credentials")

            # Check connectivity with new credentials
            return self._check_service_connectivity()

        except Exception as e:
            logger.error(f"Unexpected error in _get_new_credentials: {str(e)}")
            self.console.print(
                "[red]An unexpected error occurred while getting credentials.[/red]\n"
                "[yellow]Please try again later or contact support if the issue persists.[/yellow]"
            )
            return False
        finally:
            self._stop_progress()

    def _get_env_credentials(self) -> bool:
        """Try to get credentials from environment variables"""
        try:
            nbk_id = os.getenv(AUTH_CONFIG['NBK_ID_ENV_VAR'])
            password = os.getenv(AUTH_CONFIG['PASSWORD_ENV_VAR'])

            if nbk_id and password and self._validate_nbk_id(nbk_id):
                self.username = nbk_id.strip()  # Ensure no whitespace
                self.password = password.strip()
                logger.info("Using credentials from environment variables")
                return True

            logger.warning("Missing or invalid environment credentials")
            return False
        except Exception as e:
            logger.error(f"Error reading environment credentials: {str(e)}")
            return False

    def _is_interactive(self) -> bool:
        """Check if we're running in an interactive environment"""
        return sys.stdin.isatty()

    def authenticate(self) -> bool:
        """Securely get credentials and test authentication with visual feedback"""
        try:
            self._start_progress("[bold cyan]🔐 Setting up authentication...[/bold cyan]")
            self._update_progress("[cyan]Checking environment credentials...[/cyan]", advance=10)

            # First try environment variables
            if self._get_env_credentials():
                self.ansible_instance = cast(
                    Literal['NON_PROD', 'PROD'],
                    os.getenv('ANSIBLE_INSTANCE', AUTH_CONFIG['DEFAULT_INSTANCE'])
                )
                self._update_progress("[green]✓ Found environment credentials![/green]", advance=30)
                logger.info(f"Using {self.ansible_instance} instance from environment")
                return self._check_service_connectivity()

            elif self._is_interactive():
                self._update_progress("[cyan]Starting interactive authentication...[/cyan]", advance=20)
                self._stop_progress()  # Stop progress for user input

                # Instance selection with visual prompt
                self.console.print("\n[bold cyan]Select Ansible Instance:[/bold cyan]")
                self.ansible_instance = cast(
                    Literal['NON_PROD', 'PROD'],
                    Prompt.ask(
                        "Choose environment",
                        choices=["NON_PROD", "PROD"],
                        default="NON_PROD"
                    )
                )

                self._start_progress("[cyan]Checking saved credentials...[/cyan]")
                saved_credentials = self.credential_manager.load_credentials()
                if saved_credentials:
                    self.username, self.password = saved_credentials
                    self._update_progress("[green]✓ Using saved credentials...[/green]", advance=40)
                    logger.info("Using saved credentials for authentication")
                    return self._check_service_connectivity()
                else:
                    self._update_progress("[yellow]No saved credentials found[/yellow]", advance=30)
                    logger.info("No saved credentials found, requesting new credentials")
                    return self._get_new_credentials()
            else:
                self._update_progress("[red]✗ No credentials available[/red]", advance=100)
                logger.error("No credentials available and not in interactive mode")
                self.console.print(
                    "\n[bold red]Error: Not running interactively and no environment credentials found.[/bold red]"
                    f"\n[yellow]Please set {AUTH_CONFIG['NBK_ID_ENV_VAR']} and {AUTH_CONFIG['PASSWORD_ENV_VAR']} "
                    "environment variables.[/yellow]"
                )
                return False

        except KeyboardInterrupt:
            logger.info("Authentication cancelled by user")
            self.console.print("\n[yellow]⚠ Authentication cancelled by user.[/yellow]")
            return False
        except Exception as e:
            logger.error(f"Unexpected error during authentication: {str(e)}")
            self.console.print(
                "\n[bold red]✗ An unexpected error occurred during authentication.[/bold red]"
                "\n[yellow]Please try again or contact support if the issue persists.[/yellow]"
            )
            return False
        finally:
            self._stop_progress()

    def _check_service_connectivity(self) -> bool:
        """Check connectivity to services with enhanced visual feedback"""
        if not all([self.username, self.password, self.ansible_instance]):
            logger.error("Missing required authentication parameters")
            return False

        try:
            self._start_progress("[bold cyan]🔄 Checking service connectivity...[/bold cyan]")

            service_status = {
                'Ansible Tower': {},
                'XLR': {}
            }

            vpn_issues = False
            auth_issues = False
            other_issues = False

            if self.ansible_instance is None:
                logger.error("Ansible instance is not set")
                return False

            instance_envs = TOWER_BASE_URLS[self.ansible_instance]
            total_services = len(instance_envs) * 2  # Tower + XLR for each env
            progress_per_service = 100 // total_services

            # Check Ansible Tower connectivity
            for env, url in instance_envs.items():
                self._update_progress(
                    f"[cyan]⟳ Checking Ansible Tower {env} connectivity...[/cyan]",
                    advance=progress_per_service // 2
                )

                if not (self.username and self.password):
                    logger.error("Username or password is not set")
                    return False

                logger.debug(f"Attempting to connect to Ansible Tower {env} at URL: {url}")
                try:
                    api_url = f"{url}/me/"
                    logger.info(f"Making request to: {api_url}")
                    response = requests.get(
                        api_url,
                        auth=HTTPBasicAuth(self.username, self.password),
                        timeout=10,
                        verify=True
                    )

                    if response.status_code == 401:
                        logger.error(f"Authentication failed for {env} - Invalid credentials")
                        self.credential_manager.mark_login_failed()
                        auth_issues = True
                        service_status['Ansible Tower'][env] = False
                        continue

                    service_status['Ansible Tower'][env] = response.status_code == 200
                    status_icon = "✓" if response.status_code == 200 else "✗"
                    status_color = "green" if response.status_code == 200 else "red"
                    self._update_progress(
                        f"[{status_color}]{status_icon} Ansible Tower {env} connection: {response.status_code}[/{status_color}]",
                        advance=progress_per_service // 2
                    )

                except requests.exceptions.ConnectionError as e:
                    if "NameResolutionError" in str(e):
                        vpn_issues = True
                    logger.error(f"Connection error for Ansible Tower {env}: {str(e)}")
                    service_status['Ansible Tower'][env] = False
                except requests.exceptions.RequestException as e:
                    other_issues = True
                    logger.error(f"Failed to connect to Ansible Tower {env}: {str(e)}")
                    service_status['Ansible Tower'][env] = False

            # Add visual separator
            self.console.print("\n[dim]─" * 50)

            # Check XLR connectivity
            for env, url in XLR_CONFIG['BASE_URLS'].items():
                if env in instance_envs:
                    self._update_progress(
                        f"[cyan]⟳ Checking XLR {env} connectivity...[/cyan]",
                        advance=progress_per_service // 2
                    )
                    logger.debug(f"Attempting to connect to XLR {env} at URL: {url}")
                    try:
                        api_url = f"{url}/api/{XLR_CONFIG['API_VERSION']}/profile"
                        logger.info(f"Making request to: {api_url}")
                        response = requests.get(
                            api_url,
                            auth=HTTPBasicAuth(self.username, self.password),
                            timeout=10,
                            verify=True
                        )

                        if response.status_code == 401:
                            logger.error(f"Authentication failed for XLR {env}")
                            self.credential_manager.mark_login_failed()
                            auth_issues = True
                            service_status['XLR'][env] = False
                            continue

                        service_status['XLR'][env] = response.status_code == 200
                        status_icon = "✓" if response.status_code == 200 else "✗"
                        status_color = "green" if response.status_code == 200 else "red"
                        self._update_progress(
                            f"[{status_color}]{status_icon} XLR {env} connection: {response.status_code}[/{status_color}]",
                            advance=progress_per_service // 2
                        )

                    except requests.exceptions.ConnectionError as e:
                        if "NameResolutionError" in str(e):
                            vpn_issues = True
                        logger.error(f"Connection error for XLR {env}: {str(e)}")
                        service_status['XLR'][env] = False
                    except requests.exceptions.RequestException as e:
                        other_issues = True
                        logger.error(f"Failed to connect to XLR {env}: {str(e)}")
                        service_status['XLR'][env] = False

            ansible_connected = any(service_status['Ansible Tower'].values())
            xlr_connected = any(service_status['XLR'].values())

            if ansible_connected and xlr_connected:
                self._update_progress("[bold green]✓ Successfully connected to all services![bold green]", advance=100)
                logger.info("Successfully connected to both Ansible Tower and XLR services")
                return True

            # Generate user-friendly error messages based on issue type
            error_messages = []

            if vpn_issues:
                error_messages.append(
                    "[bold red]✗ VPN Connection Required[/bold red]\n"
                    "[yellow]It seems you're not connected to the office network. Please:\n"
                    "1. Connect to your office VPN\n"
                    "2. Ensure you can access internal services\n"
                    "3. Try running the tool again[/yellow]"
                )

            if auth_issues:
                error_messages.append(
                    "[bold red]✗ Authentication Failed[/bold red]\n"
                    "[yellow]Your credentials were not accepted. Please:\n"
                    "1. Verify your NBK ID and password\n"
                    "2. Check if your password has expired\n"
                    "3. Try again with correct credentials[/yellow]"
                )

            if other_issues and not (vpn_issues or auth_issues):
                unreachable_services = []
                for service, envs in service_status.items():
                    failed_envs = [env for env, status in envs.items() if not status]
                    if failed_envs:
                        unreachable_services.append(f"{service} ({', '.join(failed_envs)})")

                if unreachable_services:
                    error_messages.append(
                        "[bold red]✗ Service Connection Issues[/bold red]\n"
                        f"[yellow]Unable to connect to: {', '.join(unreachable_services)}\n"
                        "Please verify:\n"
                        "1. The services are currently operational\n"
                        "2. You have the necessary permissions\n"
                        "3. Try again in a few minutes[/yellow]"
                    )

            error_list = "\n\n".join(error_messages)
            self._update_progress("[red]✗ Service connectivity check failed[/red]", advance=100)
            self.console.print("\n" + error_list)
            return False

        except Exception as e:
            logger.error(f"Error checking service connectivity: {str(e)}")
            self.console.print(
                "\n[bold red]✗ Error checking service connectivity.[/bold red]"
                "\n[yellow]Please verify your network connection and try again.[/yellow]"
            )
            return False
        finally:
            self._stop_progress()

    def _validate_url(self, url: str) -> bool:
        """Validate URL format and accessibility"""
        try:
            result = urlparse(url)
            return all([result.scheme in ['http', 'https'], result.netloc])
        except Exception:
            return False

    def get_inventory_data(self, spk_name: str, environment: str) -> List[Dict[str, Any]]:
        """Fetch inventory data with enhanced visual progress indicators"""
        try:
            self._start_progress(f"[bold cyan]🔍 Searching for {environment} inventories matching '{spk_name}'...[/bold cyan]")
            inventory_search = spk_name

            if not self.ansible_instance:
                logger.error("Ansible instance is not set")
                return []

            if not all([self.username, self.password]):
                logger.error("Username or password is not set")
                return []

            base_url = TOWER_BASE_URLS[self.ansible_instance][environment]

            # Get inventories matching the SPK
            self._update_progress("[cyan]⟳ Querying Ansible Tower API...[/cyan]", advance=10)
            inventories_url = f"{base_url}/inventories/?search={inventory_search}"
            inventories_response = self.session.get(
                inventories_url,
                auth=HTTPBasicAuth(self.username, self.password),
                timeout=30
            )
            inventories_response.raise_for_status()
            inventories = inventories_response.json()['results']

            if not inventories:
                self._update_progress(
                    f"[yellow]⚠ No inventories found for '{spk_name}'[/yellow]",
                    advance=100
                )
                logger.warning(f"No inventories found for {inventory_search}")
                return []

            self._update_progress(
                f"[green]✓ Found {len(inventories)} matching inventories![/green]",
                advance=20
            )

            # Calculate progress increments
            total_inventories = len(inventories)
            progress_per_inventory = 60 // total_inventories  # Integer division for progress

            inventory_data = []
            for idx, inventory in enumerate(inventories, 1):
                self._update_progress(
                    f"[cyan]⟳ Processing inventory {idx}/{total_inventories}: {inventory['name']}[/cyan]",
                    advance=5
                )

                # Get groups for each inventory
                groups_url = f"{base_url}/inventories/{inventory['id']}/groups/"
                groups_response = self.session.get(
                    groups_url,
                    auth=HTTPBasicAuth(self.username, self.password),
                    timeout=30
                )
                groups_response.raise_for_status()
                groups = groups_response.json()['results']

                self._update_progress(
                    f"[cyan]↳ Found {len(groups)} groups in {inventory['name']}[/cyan]",
                    advance=5
                )

                # Process each group with integer progress updates
                total_groups = len(groups)
                progress_per_group = progress_per_inventory // max(total_groups, 1)

                for group_idx, group in enumerate(groups, 1):
                    self._update_progress(
                        f"[cyan]Processing group {group_idx}/{total_groups}: {group['name']}[/cyan]"
                    )

                    # Get hosts for each group
                    hosts_url = f"{base_url}/groups/{group['id']}/hosts/"
                    hosts_response = self.session.get(
                        hosts_url,
                        auth=HTTPBasicAuth(self.username, self.password),
                        timeout=30
                    )
                    hosts_response.raise_for_status()
                    hosts = hosts_response.json()['results']

                    # Add host data to inventory_data
                    for host in hosts:
                        inventory_data.append({
                            'inventory_name': inventory['name'],
                            'group': group['name'],
                            'host_fqdn': host['name'],
                            'is_enabled': host['enabled'],
                            'inventory_id': inventory['id']
                        })

                    self._update_progress(
                        f"[green]✓ Found {len(hosts)} hosts in group {group['name']}[/green]",
                        advance=progress_per_group
                    )

                self._update_progress(
                    f"[green]✓ Completed processing inventory: {inventory['name']}[/green]"
                )

            final_count = len(inventory_data)
            status_color = "green" if final_count > 0 else "yellow"
            status_icon = "✓" if final_count > 0 else "⚠"
            self._update_progress(
                f"[{status_color}]{status_icon} Retrieved {final_count} total inventory items for {environment}![/{status_color}]",
                advance=40
            )
            return inventory_data

        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching inventory data: {str(e)}")
            self.console.print(f"[bold red]✗ Error fetching inventory data: {str(e)}[/bold red]")
            return []
        finally:
            self._stop_progress()

    def generate_excel_report(self, spk_name: str, all_inventory_data: List[Dict[str, Any]]) -> Optional[str]:
        """Generate Excel report with the inventory data"""
        try:
            self._start_progress("[cyan]Generating Excel report...[/cyan]")
            if not all_inventory_data:
                logger.warning("No data to generate report")
                return None

            # Initialize workbook
            self._update_progress("[cyan]Creating workbook...[/cyan]")
            wb = Workbook()
            ws = cast(Worksheet, wb.active)
            if ws is None:
                raise ValueError("Failed to get active worksheet")

            ws.title = "Inventory Report"

            # Add headers with styling
            self._update_progress("[cyan]Adding headers...[/cyan]")
            headers = ['Inventory Name', 'Groups', 'Server/Host FQDN', 'IsEnabled', 'Inventory ID']
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = header_alignment

            # Add data
            self._update_progress("[cyan]Adding inventory data...[/cyan]")
            for row_idx, item in enumerate(all_inventory_data, start=2):
                for col_idx, value in enumerate([
                    item['inventory_name'],
                    item['group'],
                    item['host_fqdn'],
                    item['is_enabled'],
                    item['inventory_id']
                ], start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='left')

            # Format the worksheet
            self._update_progress("[cyan]Formatting worksheet...[/cyan]")
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 25

            # Save the file
            self._update_progress("[cyan]Saving Excel file...[/cyan]")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if self.output_dir is None:
                raise ValueError("Output directory is not accessible")

            filename = os.path.join(self.output_dir, f"ansible_inventory_{sanitize_filename(spk_name)}_{timestamp}.xlsx")

            try:
                wb.save(filename)
                self._update_progress("[green]Excel report generated successfully![/green]")
                return filename
            except PermissionError:
                logger.error("Permission denied when saving the file")
                self._update_progress("[red]Error: Unable to save the file. Please check permissions.[/red]")
                return None
            except Exception as e:
                logger.error(f"Error saving Excel file: {str(e)}")
                self._update_progress("[red]Error: Unable to save the Excel file.[/red]")
                return None

        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            self.console.print(f"[red]Error generating Excel report: {str(e)}[/red]")
            return None
        finally:
            self._stop_progress()

    def _get_xlr_components(self, environment: str) -> List[str]:
        """Get component names from XLR train"""
        try:
            self._start_progress("[cyan]Fetching components from XLR train...[/cyan]")
            xlr_train_url = Prompt.ask("Enter the XLR train URL")

            if not self.xlr_client:
                if not all([self.username, self.password]):
                    logger.error("Username and password are not set.")
                    return []
                self.xlr_client = XLRClient(environment, self.username, self.password)

            components = self.xlr_client.get_components_from_train(xlr_train_url)
            self._update_progress("[green]Fetching components from XLR train...[/green]")
            return components
        except Exception as e:
            logger.error(f"Error fetching XLR components: {str(e)}")
            self.console.print(
                "[red]Error fetching components from XLR train.[/red]\n"
                "[yellow]Please verify:\n"
                "1. The XLR train URL is correct\n"
                "2. You have access to the XLR train\n"
                "3. The train contains components in the Relationship view[/yellow]"
            )
            return []
        finally:
            self._stop_progress()

    def run(self):
        """Main execution flow"""
        try:
            self._start_progress("[cyan]Initializing Ansible Tower Inventory Reporter...[/cyan]")

            self.console.print("[blue]Ansible Tower Inventory Reporter[/blue]")
            self._update_progress("[cyan]Setting up authentication...[/cyan]")

            if not self.authenticate():
                self.console.print("[red]Authentication failed. Please try again.[/red]")
                return

            while True:
                # Prompt for SPK with styling
                self.console.print("\n[cyan]SPK Information[/cyan]")
                spk_name = Prompt.ask(
                    "[cyan]Enter the SPK name[/cyan] (e.g., ASAPREQ)",
                    show_default=True
                )

                if not validate_spk_name(spk_name):
                    self.console.print(
                        "[red]Invalid SPK name format.[/red]\n"
                        "[yellow]SPK name should:\n"
                        "• Be 3-50 characters long\n"
                        "• Contain only letters, numbers, underscores, and hyphens[/yellow]"
                    )
                    continue

                # First get components from XLR - use appropriate environment based on Ansible instance
                xlr_env = 'PROD' if self.ansible_instance == 'PROD' else 'DEV'

                self._update_progress("[cyan]Fetching XLR components...[/cyan]")
                components = self._get_xlr_components(xlr_env)
                self._update_progress("[green]XLR component fetch complete![/green]")

                if not components:
                    if not Confirm.ask(
                        "[yellow]No components found in XLR train. Would you like to continue with manual SPK search?[/yellow]"
                    ):
                        return

                all_inventory_data = []
                envs = self.environments[self.ansible_instance]

                for env in envs:
                    self._update_progress(f"[cyan]Processing {env} environment...[/cyan]")

                    # If we have components from XLR, search for each component
                    if components:
                        for i, component in enumerate(components, 1):
                            self._update_progress(
                                f"[cyan]Fetching {env} inventory for component {i}/{len(components)}: {component}[/cyan]"
                            )
                            inventory_data = self.get_inventory_data(f"{spk_name}_{component}", env)
                            all_inventory_data.extend(inventory_data)
                    else:
                        # Fallback to regular SPK search
                        self._update_progress(f"[cyan]Fetching {env} inventory for {spk_name}...[/cyan]")
                        inventory_data = self.get_inventory_data(spk_name, env)
                        all_inventory_data.extend(inventory_data)

                    self._update_progress(f"[green]Completed {env} inventory retrieval![/green]")

                if all_inventory_data:
                    self._update_progress("[cyan]Generating Excel report...[/cyan]")
                    filename = self.generate_excel_report(spk_name, all_inventory_data)
                    if filename:
                        self._update_progress("[green]Report generated successfully![/green]")
                        self.console.print(f"\n[green]Report generated successfully: {filename}[/green]")

                else:
                    self.console.print(
                        f"[yellow]No inventory data found for the specified SPK in {self.ansible_instance} environments.[/yellow]"
                    )

                if not Confirm.ask("[cyan]Would you like to generate another report?[/cyan]"):
                    break

        except KeyboardInterrupt:
            self.console.print("\n[yellow]Program terminated by user.[/yellow]")
        except Exception as e:
            logger.error(f"Unexpected error in run(): {str(e)}")
            self.console.print(
                "[red]An unexpected error occurred.[/red]\n"
                "[yellow]Please check the logs for details.[/yellow]"
            )
        finally:
            self._stop_progress()

if __name__ == "__main__":
    try:
        reporter = AnsibleTowerInventoryReporter()
        reporter.run()
    except KeyboardInterrupt:
        console.print("\n[yellow]Program terminated by user.[/yellow]")
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        console.print("[red]An unexpected error occurred. Please check the logs for details.[/red]")